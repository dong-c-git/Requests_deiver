# coding:utf-8

from openpyxl import load_workbook
import openpyxl
import os,time

def copy_excel(excelpath1,excelpath2):
    """复制excel,把excelpath1的数据复制到excelpath2"""

    wb2 = openpyxl.Workbook(excelpath2)
    wb2.save(excelpath2)  # 保存数据
    #读取数据
    wb1 = openpyxl.load_workbook(excelpath1)
    wb2 = openpyxl.load_workbook(excelpath2)
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames
    sheet1 = wb1[sheets1[0]]
    sheet2 = wb2[sheets2[0]]
    max_row = sheet1.max_row          #最大行数
    max_column = sheet1.max_column    #最大列数
    for m in list(range(1,max_row+1)):
        for n in list(range(97,97+max_column)):          #char(97)='a'
            n = chr(n)
            i = "%s%d"%(n,m)                             #单元格编码
            cell1 = sheet1[i].value                      #获取data单元格数据
            sheet2[i].value = cell1                      #赋值到test单元格
    wb2.save(excelpath2)                                #保存数据
    wb1.close()                                          #关闭excel
    wb2.close()

class Write_excel(object):
    """修改excel数据"""

    def __init__(self,filename):
        self.filename = filename
        self.wd = openpyxl.load_workbook(self.filename)
        self.ws = self.wd.active    #激活sheet

    def write(self,row_n,col_n,value):
        """写入数据，如（2,3，'hello world'）第二行第三列数据写入hello world"""
        self.ws.cell(row_n,col_n).value = value
        self.wd.save(self.filename)

if __name__=="__main__":
    excelpath1="debug_api.xlsx"
    excelpath2="debug_api2.xlsx"
    copy_excel(excelpath1,excelpath2)
    wt = Write_excel("debug_api2.xlsx")
    wt.write(4,5,"helleop")
    wt.write(5,5,"helleop")




